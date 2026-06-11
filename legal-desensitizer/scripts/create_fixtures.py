"""Generate synthetic test fixtures for 003-document-io tests."""

import os
import zipfile
from lxml import etree

FIXTURES_DIR = os.path.join(os.path.dirname(__file__), "..", "tests", "fixtures")


def _ensure_dir(path):
    os.makedirs(path, exist_ok=True)


def create_docx_fixtures():
    """Create DOCX fixtures with cross-run entity scenarios."""
    _ensure_dir(os.path.join(FIXTURES_DIR, "docx"))

    _create_docx(
        os.path.join(FIXTURES_DIR, "docx", "sample.docx"),
        paragraphs=[
            "张三的手机号是13800138000，邮箱是zhangsan@example.com。",
            "李四的案号为(2024)京0101民初12345号。",
        ],
    )

    _create_docx_with_runs(
        os.path.join(FIXTURES_DIR, "docx", "cross_run.docx"),
        paragraphs_runs=[
            [("联系", None), ("电话13800138000请拨打。", None)],
            [("案号(2024)京0101民初12345号已登记。", None)],
        ],
    )

    _create_docx(
        os.path.join(FIXTURES_DIR, "docx", "cross_paragraph.docx"),
        paragraphs=[
            "请联系1380013",
            "8000获取详情。",
        ],
    )

    _create_docx(
        os.path.join(FIXTURES_DIR, "docx", "no_match.docx"),
        paragraphs=["这是一份没有任何敏感信息的文档。", "第二段也是安全的。"],
    )

    _create_docx(
        os.path.join(FIXTURES_DIR, "docx", "empty.docx"),
        paragraphs=[],
    )


def create_xlsx_fixtures():
    """Create XLSX fixtures using openpyxl."""
    _ensure_dir(os.path.join(FIXTURES_DIR, "xlsx"))

    import openpyxl

    # ── sample.xlsx ──
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "姓名"
    ws["B1"] = "电话"
    ws["C1"] = "邮箱"
    ws["A2"] = "张三"
    ws["B2"] = "13800138000"
    ws["C2"] = "zhangsan@example.com"
    ws["A3"] = "李四"
    ws["B3"] = "13900139000"
    ws["C3"] = "lisi@company.cn"
    wb.save(os.path.join(FIXTURES_DIR, "xlsx", "sample.xlsx"))

    # ── shared_strings.xlsx ──
    # Both A2 and A3 contain "张三" - openpyxl will use shared strings
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "人员表"
    ws["A1"] = "姓名"
    ws["B1"] = "电话"
    ws["A2"] = "张三"
    ws["B2"] = "13800138000"
    ws["A3"] = "张三"  # Same value as A2 - tests shared string safety
    ws["B3"] = "13900139000"
    wb.save(os.path.join(FIXTURES_DIR, "xlsx", "shared_strings.xlsx"))

    # ── formula.xlsx ──
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "项目"
    ws["B1"] = "金额"
    ws["A2"] = "合同款"
    ws["B2"] = 10000
    ws["A3"] = "税费"
    ws["B3"] = "=B2*0.06"  # Formula cell
    ws["A4"] = "联系人"
    ws["B4"] = "13800138000"
    wb.save(os.path.join(FIXTURES_DIR, "xlsx", "formula.xlsx"))

    # ── empty_cells.xlsx ──
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "张三"
    ws["B1"] = ""  # Empty cell
    ws["C1"] = "13800138000"
    ws["A2"] = ""  # Empty cell
    ws["B2"] = "李四"
    ws["C2"] = ""  # Empty cell
    wb.save(os.path.join(FIXTURES_DIR, "xlsx", "empty_cells.xlsx"))

    # ── multi_shared.xlsx: explicit shared string test ──
    # Create by manually writing XML to ensure shared strings are actually shared
    _create_xlsx_with_shared_strings(
        os.path.join(FIXTURES_DIR, "xlsx", "multi_shared.xlsx"),
        shared_texts=["姓名", "电话", "张三", "13800138000", "13900139000"],
        cells=[
            # sheet, row, col, ss_index or None, formula
            (0, 1, 1, 0, None),   # A1 = "姓名"
            (0, 1, 2, 1, None),   # B1 = "电话"
            (0, 2, 1, 2, None),   # A2 = "张三"
            (0, 2, 2, 3, None),   # B2 = "13800138000"
            (0, 3, 1, 2, None),   # A3 = "张三" (same ss_index=2 as A2)
            (0, 3, 2, 4, None),   # B3 = "13900139000"
        ],
        sheet_names=["Sheet1"],
    )


def create_pdf_fixtures():
    """Create a simple PDF fixture for audit testing."""
    _ensure_dir(os.path.join(FIXTURES_DIR, "pdf"))

    pdf_path = os.path.join(FIXTURES_DIR, "pdf", "sample.pdf")
    _create_minimal_pdf(pdf_path, "张三的手机号是13800138000。\n邮箱: zhangsan@example.com\n")


def _create_docx(path, paragraphs):
    """Create a minimal DOCX file with given paragraphs."""
    _write_docx_xml(path, paragraphs, runs_per_paragraph=None)


def _create_docx_with_runs(path, paragraphs_runs):
    """Create a DOCX where each paragraph has specific runs."""
    _write_docx_xml(path, None, runs_per_paragraph=paragraphs_runs)


def _write_docx_xml(path, paragraphs=None, runs_per_paragraph=None):
    """Write a minimal DOCX ZIP file."""
    W_NS = "http://schemas.openxmlformats.org/wordprocessingml/2006/main"
    R_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    CT_NS = "http://schemas.openxmlformats.org/package/2006/content-types"
    RELS_NS = "http://schemas.openxmlformats.org/package/2006/relationships"

    body_children = []

    if paragraphs is not None:
        for para_text in paragraphs:
            p = etree.Element(f"{{{W_NS}}}p")
            if para_text:
                r = etree.SubElement(p, f"{{{W_NS}}}r")
                t = etree.SubElement(r, f"{{{W_NS}}}t")
                t.text = para_text
                t.set("{http://www.w3.org/XML/1998/namespace}space", "preserve")
            body_children.append(p)
    elif runs_per_paragraph is not None:
        for runs in runs_per_paragraph:
            p = etree.Element(f"{{{W_NS}}}p")
            for run_text, _ in runs:
                r = etree.SubElement(p, f"{{{W_NS}}}r")
                t = etree.SubElement(r, f"{{{W_NS}}}t")
                t.text = run_text
                t.set("{http://www.w3.org/XML/1998/namespace}space", "preserve")
            body_children.append(p)

    doc = etree.Element(f"{{{W_NS}}}document", nsmap={"w": W_NS, "r": R_NS})
    body = etree.SubElement(doc, f"{{{W_NS}}}body")
    for child in body_children:
        body.append(child)

    doc_xml = etree.tostring(doc, xml_declaration=True, encoding="UTF-8", standalone=True)

    rels_xml = f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="{RELS_NS}">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" target="word/document.xml"/>
</Relationships>"""

    ct_xml = f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Types xmlns="{CT_NS}">
  <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>
  <Default Extension="xml" ContentType="application/xml"/>
  <Override PartName="/word/document.xml" ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>
</Types>"""

    word_rels_xml = f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="{RELS_NS}">
</Relationships>"""

    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("[Content_Types].xml", ct_xml)
        zf.writestr("_rels/.rels", rels_xml)
        zf.writestr("word/document.xml", doc_xml)
        zf.writestr("word/_rels/document.xml.rels", word_rels_xml)


def _col_letter(idx):
    """0-based index to column letter: 0->A, 1->B, etc."""
    result = ""
    idx += 1
    while idx > 0:
        idx, rem = divmod(idx - 1, 26)
        result = chr(rem + ord("A")) + result
    return result


def _create_xlsx_with_shared_strings(path, shared_texts, cells, sheet_names):
    """Create an XLSX with explicit shared string references."""
    SS_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    RELS_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
    CT_NS = "http://schemas.openxmlformats.org/package/2006/content-types"
    XML_NS = "http://www.w3.org/XML/1998/namespace"
    R_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"

    # Build sharedStrings.xml manually
    ss_lines = [
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>',
        f'<sst xmlns="{SS_NS}" count="{len(shared_texts)}" uniqueCount="{len(shared_texts)}">',
    ]
    for text in shared_texts:
        ss_lines.append(f"  <si><t xml:space=\"preserve\">{text}</t></si>")
    ss_lines.append("</sst>")
    ss_xml = "\n".join(ss_lines).encode("utf-8")

    # Build sheet XMLs
    sheets_data = {}
    for sheet_idx, row, col, ss_idx, formula in cells:
        if sheet_idx not in sheets_data:
            sheets_data[sheet_idx] = []
        sheets_data[sheet_idx].append((row, col, ss_idx, formula))

    sheet_xmls = []
    for sheet_idx in sorted(sheets_data.keys()):
        rows = {}
        for row, col, ss_idx, formula in sheets_data[sheet_idx]:
            if row not in rows:
                rows[row] = []
            rows[row].append((col, ss_idx, formula))

        lines = [
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>',
            f'<worksheet xmlns="{SS_NS}">',
            "  <sheetData>",
        ]
        for row_num in sorted(rows.keys()):
            lines.append(f'    <row r="{row_num}">')
            for col, ss_idx, formula in rows[row_num]:
                ref = f"{_col_letter(col - 1)}{row_num}"
                if formula:
                    lines.append(f'      <c r="{ref}"><f>{formula}</f></c>')
                elif ss_idx is not None:
                    lines.append(f'      <c r="{ref}" t="s"><v>{ss_idx}</v></c>')
                else:
                    lines.append(f'      <c r="{ref}"/>')
            lines.append("    </row>")
        lines.append("  </sheetData>")
        lines.append("</worksheet>")
        sheet_xmls.append("\n".join(lines).encode("utf-8"))

    # Build workbook.xml
    wb_lines = [
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>',
        f'<workbook xmlns="{SS_NS}" xmlns:r="{R_NS}">',
        "  <sheets>",
    ]
    for i, name in enumerate(sheet_names):
        wb_lines.append(f'    <sheet name="{name}" sheetId="{i+1}" r:id="rId{i+1}"/>')
    wb_lines.append("  </sheets>")
    wb_lines.append("</workbook>")
    wb_xml = "\n".join(wb_lines).encode("utf-8")

    # Relationships
    rels_xml = f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<Relationships xmlns="{RELS_NS}">
  <Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" target="xl/workbook.xml"/>
</Relationships>""".encode()

    wb_rels_lines = [
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>',
        f'<Relationships xmlns="{RELS_NS}">',
    ]
    for i in range(len(sheet_xmls)):
        wb_rels_lines.append(f'  <Relationship Id="rId{i+1}" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" target="worksheets/sheet{i+1}.xml"/>')
    wb_rels_lines.append(f'  <Relationship Id="rId99" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/sharedStrings" target="sharedStrings.xml"/>')
    wb_rels_lines.append("</Relationships>")
    wb_rels_xml = "\n".join(wb_rels_lines).encode()

    ct_lines = [
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>',
        f'<Types xmlns="{CT_NS}">',
        '  <Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>',
        '  <Default Extension="xml" ContentType="application/xml"/>',
        '  <Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/>',
        '  <Override PartName="/xl/sharedStrings.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml"/>',
    ]
    for i in range(len(sheet_xmls)):
        ct_lines.append(f'  <Override PartName="/xl/worksheets/sheet{i+1}.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/>')
    ct_lines.append("</Types>")
    ct_xml = "\n".join(ct_lines).encode()

    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("[Content_Types].xml", ct_xml)
        zf.writestr("_rels/.rels", rels_xml)
        zf.writestr("xl/workbook.xml", wb_xml)
        zf.writestr("xl/_rels/workbook.xml.rels", wb_rels_xml)
        zf.writestr("xl/sharedStrings.xml", ss_xml)
        for i, sheet_xml in enumerate(sheet_xmls):
            zf.writestr(f"xl/worksheets/sheet{i+1}.xml", sheet_xml)


def _create_minimal_pdf(path, text):
    """Create a minimal PDF with extractable text."""
    lines = []
    objects = []

    lines.append(b"%PDF-1.4")

    obj1_offset = sum(len(l) + 1 for l in lines)
    objects.append((1, obj1_offset))
    lines.append(b"1 0 obj<</Type/Catalog/Pages 2 0 R>>endobj")

    obj2_offset = sum(len(l) + 1 for l in lines)
    objects.append((2, obj2_offset))
    lines.append(b"2 0 obj<</Type/Pages/Kids[3 0 R]/Count 1>>endobj")

    obj3_offset = sum(len(l) + 1 for l in lines)
    objects.append((3, obj3_offset))
    lines.append(b"3 0 obj<</Type/Page/Parent 2 0 R/MediaBox[0 0 612 792]/Contents 4 0 R/Resources<</Font<</F1 5 0 R>>>>>>endobj")

    encoded_text = text.encode("latin-1", errors="replace")
    stream_content = f"BT /F1 12 Tf 100 700 Td ({encoded_text.decode('latin-1')}) Tj ET"
    stream_bytes = stream_content.encode("latin-1")
    obj4_offset = sum(len(l) + 1 for l in lines)
    objects.append((4, obj4_offset))
    lines.append(f"4 0 obj<</Length {len(stream_bytes)}>>stream".encode())
    lines.append(stream_bytes)
    lines.append(b"endstream endobj")

    obj5_offset = sum(len(l) + 1 for l in lines)
    objects.append((5, obj5_offset))
    lines.append(b"5 0 obj<</Type/Font/Subtype/Type1/BaseFont/Helvetica>>endobj")

    xref_offset = sum(len(l) + 1 for l in lines)
    lines.append(b"xref")
    lines.append(f"0 {len(objects) + 1}".encode())
    lines.append(b"0000000000 65535 f ")
    for obj_num, offset in objects:
        lines.append(f"{offset:010d} 00000 n ".encode())

    lines.append(b"trailer")
    lines.append(f"<</Size {len(objects) + 1}/Root 1 0 R>>".encode())
    lines.append(b"startxref")
    lines.append(f"{xref_offset}".encode())
    lines.append(b"%%EOF")

    with open(path, "wb") as f:
        for line in lines:
            f.write(line + b"\n")


if __name__ == "__main__":
    create_docx_fixtures()
    create_xlsx_fixtures()
    create_pdf_fixtures()
    print("Fixtures created successfully.")
