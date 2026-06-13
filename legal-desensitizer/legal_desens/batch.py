"""Batch case redaction orchestrator (018).

Thin orchestration layer that composes existing capabilities:
profile / allow-deny / redact / redact-scan / report / gate.
No new desensitization logic.
"""

from __future__ import annotations

import hashlib
import json
import os
import re
import shutil
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple

from .audit import audit
from .engine.allowlist import load_allowlist
from .engine.ner import inspect_ner
from .io import read_text, sha256_file
from .profile import load_profile, resolve_profile_name
from .redact import redact
from .rules import load_rules
from .scan import scan_redact

# ── Format routing (mirrors cli.py decision table) ──────────────────────────

_FORMAT_A = {".txt", ".md", ".csv", ".docx", ".xlsx"}
_FORMAT_B = {".pdf", ".png", ".jpg", ".jpeg", ".tiff", ".bmp"}
_FORMAT_C = {".doc", ".xls", ".wps", ".et", ".dps", ".pages", ".numbers", ".key"}

_C_CONVERT_MSG = (
    "Unsupported format: {ext}. "
    "Please convert to a supported format first:\n"
    "  - .doc -> .docx\n  - .xls -> .xlsx\n"
    "  - .wps/.et/.dps -> .docx/.xlsx/.pptx or PDF/image\n"
    "  - .pages/.numbers/.key -> .docx/.xlsx/.pptx or PDF/image\n"
    "Then run the command again on the converted file."
)

# Scan formats that require [ocr] extra
_SCAN_NEEDS_OCR = {".pdf", ".png", ".jpg", ".jpeg", ".tiff", ".bmp"}

_REPORT_FILENAME = "SENSITIVE_REDACTION_REPORT_DO_NOT_UPLOAD.md"
_SOURCE_INDEX_FILENAME = "SENSITIVE_SOURCE_INDEX_DO_NOT_UPLOAD.json"
_MANIFEST_FILENAME = "run_manifest.json"
_ARCHIVE_DIR = "_archive_sensitive_do_not_upload"
_FINAL_DIR = "final_redacted_md"
_STAGING_FINAL_DIR = "_staging_final_redacted_md"

_REPORT_FIRST_LINE = "本报告包含替换前原文，仅限本地复核，不得上传或外发。"

# Old label patterns that should NOT appear (should be 【】unnumbered)
_OLD_LABEL_PATTERN = re.compile(r"(?:人物|机构|地点|电话|银行账号|银行信息|金额|地址|身份证号)\d+")
_PRESERVED_LABELS = {
    "TIME": "【时间】",
    "MONEY": "【金额】",
}


class BatchError(Exception):
    """Raised when batch-redact-case encounters a fatal error."""


# ── Helpers ─────────────────────────────────────────────────────────────────


def _sha256_file(path: str) -> str:
    return sha256_file(path)


def _discover_files(input_dir: Path) -> List[Path]:
    """Recursively discover files in input directory."""
    files = []
    for root, _, filenames in os.walk(input_dir):
        for fn in sorted(filenames):
            fp = Path(root) / fn
            if fp.is_file():
                files.append(fp)
    return files


def _classify_ext(path: Path) -> str:
    """Classify file extension into format category."""
    ext = path.suffix.lower()
    if ext in _FORMAT_A:
        return "A"
    elif ext in _FORMAT_B:
        return "B"
    elif ext in _FORMAT_C:
        return "C"
    else:
        return "unknown"


def _check_ocr_available() -> None:
    """Check if OCR extra is installed. Raise BatchError if not."""
    try:
        from .engine.ocr import run_rapidocr  # noqa: F401
    except ImportError:
        raise BatchError(
            "OCR extra [ocr] is not installed but scan files were found. "
            "Install with: pip install legal-desens[ocr]"
        )


def _load_denylist(path: Optional[str]) -> Set[str]:
    """Load denylist from file."""
    result: Set[str] = set()
    if path:
        try:
            with open(path, "r", encoding="utf-8") as f:
                for line in f:
                    term = line.strip()
                    if term and not line.startswith("#"):
                        result.add(term)
        except FileNotFoundError:
            raise BatchError(f"Denylist file not found: {path}")
    return result


def _count_pages(path: Path, ext: str) -> int:
    """Estimate page count. Returns 1 for text formats."""
    if ext in (".docx",):
        try:
            from docx import Document
            doc = Document(str(path))
            # Rough estimate: count paragraphs / 40
            return max(1, len(doc.paragraphs) // 40 + 1)
        except Exception:
            return 1
    elif ext in (".xlsx",):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(str(path), read_only=True)
            count = len(wb.sheetnames)
            wb.close()
            return count
        except Exception:
            return 1
    else:
        return 1


# ── NER pre-check ───────────────────────────────────────────────────────────


def _precheck_ner(model_dir: Optional[str] = None) -> dict:
    """Run ner-inspect. Raises BatchError on failure (no fallback to regex-only)."""
    try:
        info = inspect_ner(model_dir)
    except (FileNotFoundError, RuntimeError, OSError) as e:
        raise BatchError(
            f"NER model inspection failed: {e}. "
            "batch-redact-case requires NER. Install or configure the model first."
        )
    return info


# ── Per-file processing ────────────────────────────────────────────────────


def _process_one(
    src: Path,
    doc_id: str,
    rules,
    profile,
    allowlist: Set[str],
    denylist: Set[str],
    model_dir: Optional[str],
) -> dict:
    """Process a single file. Returns a result dict with redacted text + metadata.

    Raises BatchError on unsupported/missing-OCR formats.
    """
    ext = src.suffix.lower()
    category = _classify_ext(src)

    if category == "C":
        raise BatchError(_C_CONVERT_MSG.format(ext=ext))

    if category == "unknown":
        raise BatchError(f"Unknown file format: {ext}. Cannot process {src.name}")

    result = {
        "doc_id": doc_id,
        "source_path": str(src),
        "source_name": src.name,
        "source_sha256": _sha256_file(str(src)),
        "source_ext": ext,
        "page_count": _count_pages(src, ext),
        "pipeline": "redact",  # or "scan"
        "irreversible": False,
        "redacted_text": None,
        "map_data": None,
        "audit_data": None,
    }

    if category == "A":
        if ext in (".txt", ".md"):
            tf = read_text(str(src))
            redacted_text, map_data, audit_data = redact(
                text=tf.text,
                rules=rules,
                source_sha256=tf.sha256,
                mode="regex+ner",
                level="labor",
                model_dir=model_dir,
                profile=profile,
                allowlist=allowlist,
                denylist=denylist,
            )
            result["redacted_text"] = redacted_text
            result["map_data"] = map_data
            result["audit_data"] = audit_data

        elif ext == ".csv":
            from .adapters.csv_adapter import csv_redact
            import tempfile
            tmp = tempfile.NamedTemporaryFile(suffix=".csv", delete=False, mode="w")
            tmp.close()

            def _txt_redact_fn(text, rules, source_sha256, mode, level, model_dir):
                return redact(
                    text, rules, source_sha256, mode, level, model_dir,
                    profile=profile, allowlist=allowlist, denylist=denylist,
                )

            try:
                map_data, audit_data = csv_redact(
                    source_path=str(src),
                    redacted_path=tmp.name,
                    redact_fn=_txt_redact_fn,
                    rules=rules,
                    mode="regex+ner",
                    level="labor",
                    model_dir=model_dir,
                )
                redacted_text = Path(tmp.name).read_text(encoding="utf-8")
                result["redacted_text"] = redacted_text
                result["map_data"] = map_data
                result["audit_data"] = audit_data
            finally:
                Path(tmp.name).unlink(missing_ok=True)

        elif ext in (".docx", ".xlsx"):
            from .adapters.docx_adapter import DOCXAdapter
            from .adapters.xlsx_adapter import XLSXAdapter
            import tempfile

            adapter = DOCXAdapter() if ext == ".docx" else XLSXAdapter()
            tmp = tempfile.NamedTemporaryFile(suffix=ext, delete=False)
            tmp.close()

            def _txt_redact_fn(text, rules, source_sha256, mode, level, model_dir):
                return redact(
                    text, rules, source_sha256, mode, level, model_dir,
                    profile=profile, allowlist=allowlist, denylist=denylist,
                )

            try:
                map_data, audit_data = adapter.redact(
                    source_path=str(src),
                    redacted_path=tmp.name,
                    redact_fn=_txt_redact_fn,
                    rules=rules,
                    mode="regex+ner",
                    level="labor",
                    model_dir=model_dir,
                )
                # For docx/xlsx, extract redacted text for report
                full_text, _ = adapter.extract_text(tmp.name)
                result["redacted_text"] = full_text
                result["map_data"] = map_data
                result["audit_data"] = audit_data
            finally:
                Path(tmp.name).unlink(missing_ok=True)

    elif category == "B":
        # Scan pipeline: requires [ocr]
        _check_ocr_available()
        result["pipeline"] = "scan"
        result["irreversible"] = True

        redacted_text, map_data, audit_data, ocr_meta = scan_redact(
            image_path=str(src),
            rules=rules,
            ocr_engine="rapidocr",
            mode="regex+ner",
            level="labor",
            model_dir=model_dir,
            profile=profile,
            allowlist=allowlist,
            denylist=denylist,
        )
        result["redacted_text"] = redacted_text
        result["map_data"] = map_data
        result["audit_data"] = audit_data
        result["ocr_meta"] = ocr_meta

    return result


# ── Validation gate ─────────────────────────────────────────────────────────


def _run_gate(
    results: List[dict],
    rules,
    profile,
    denylist: Set[str],
) -> Tuple[bool, List[str]]:
    """Run validation gate on all processed files.

    Returns (passed, failure_reasons).
    Gate checks:
    1. All files residual_scan passed
    2. No old label patterns (人物1 / 机构1)
    3. Time/money preserved (labor profile)
    4. Denylist terms no residual
    5. Manual_review / suspicious warnings displayed
    6. Bare ORG warnings filtered by allowlist
    """
    passed = True
    failures: List[str] = []

    for r in results:
        doc_id = r["doc_id"]
        text = r.get("redacted_text", "")
        audit_data = r.get("audit_data", {})
        if not text or not audit_data:
            continue

        # 1. Residual scan
        residual = audit_data.get("residual_scan", {})
        if not residual.get("passed", True):
            passed = False
            findings = residual.get("findings", [])
            failures.append(
                f"[{doc_id}] residual scan failed: {len(findings)} findings "
                f"(e.g. {findings[0].get('text_preview', '')}...)"
            )

        # 2. Old label patterns
        old_labels = _OLD_LABEL_PATTERN.findall(text)
        if old_labels:
            passed = False
            failures.append(
                f"[{doc_id}] old label pattern detected: {old_labels[:5]}"
            )

        # 3. Labor-style profiles preserve time and money. If either appears
        # in the map as redacted, or as a bracket label in output, analysis
        # signal was destroyed and the run must not be marked final.
        preserved_types = [t for t in ("TIME", "MONEY") if not profile.should_redact(t)]
        entities = (r.get("map_data") or {}).get("entities", [])
        for entity in entities:
            entity_type = entity.get("entity_type")
            if entity_type in preserved_types:
                passed = False
                failures.append(
                    f"[{doc_id}] preserved type was redacted: {entity_type}"
                )

        for entity_type in preserved_types:
            label = _PRESERVED_LABELS.get(entity_type)
            if label and label in text:
                passed = False
                failures.append(
                    f"[{doc_id}] preserved label detected in output: {label}"
                )

        # 4. Denylist residual
        text_lower = text.lower()
        for term in denylist:
            if term.lower() in text_lower:
                passed = False
                failures.append(
                    f"[{doc_id}] denylist term still present: '{term}'"
                )

        # 5. Warnings summary
        warnings = audit_data.get("warnings", [])
        bare_org = [w for w in warnings if w.get("type") == "org_bare_short_word"]
        manual_review = [w for w in warnings if w.get("type", "").startswith("manual_review")]

        if bare_org:
            # This is informational, not a gate failure (allowlist already filtered)
            pass
        if manual_review:
            # Informational: display but don't fail gate
            pass

    return passed, failures


# ── Report generation ───────────────────────────────────────────────────────


def _md_cell(value) -> str:
    """Escape a value for a compact Markdown table cell."""
    if value is None:
        return ""
    text = str(value)
    return text.replace("\n", " ").replace("|", "\\|").strip()


def _warning_text(warning: dict) -> str:
    return (
        warning.get("text")
        or warning.get("text_preview")
        or warning.get("original")
        or ""
    )


def _replacement_rows(results: List[dict]) -> List[Tuple[str, str, str, str, int, str]]:
    """Build sensitive replacement/review rows for the local-only report."""
    rows: List[Tuple[str, str, str, str, int, str]] = []

    for r in results:
        doc_id = r["doc_id"]
        map_data = r.get("map_data") or {}
        audit_data = r.get("audit_data") or {}
        warnings = audit_data.get("warnings", [])

        occurrence_counts: Dict[str, int] = {}
        for occurrence in map_data.get("occurrences", []):
            entity_id = occurrence.get("entity_id")
            if entity_id:
                occurrence_counts[entity_id] = occurrence_counts.get(entity_id, 0) + 1

        for entity in map_data.get("entities", []):
            entity_id = entity.get("id") or entity.get("entity_id") or ""
            rows.append((
                doc_id,
                entity.get("entity_type", ""),
                entity.get("original", ""),
                entity.get("replacement", ""),
                occurrence_counts.get(entity_id, 1),
                "保留脱敏",
            ))

        for warning in warnings:
            suggested_action = warning.get("suggested_action")
            if not suggested_action:
                continue
            rows.append((
                doc_id,
                warning.get("entity_type") or warning.get("type", "warning"),
                _warning_text(warning),
                "未替换",
                0,
                suggested_action,
            ))

    return rows


def _generate_report(
    results: List[dict],
    out_dir: Path,
    profile_name: str,
    allowlist_count: int,
    denylist_count: int,
    gate_passed: bool,
    gate_failures: List[str],
) -> Path:
    """Generate SENSITIVE_REDACTION_REPORT_DO_NOT_UPLOAD.md."""
    lines = [
        _REPORT_FIRST_LINE,
        "",
        f"# 脱敏报告",
        "",
        f"- **Profile**: {profile_name}",
        f"- **Allowlist 条数**: {allowlist_count}",
        f"- **Denylist 条数**: {denylist_count}",
        f"- **文档数**: {len(results)}",
        f"- **闸门**: {'通过' if gate_passed else '未通过'}",
        f"- **时间**: {datetime.now(timezone.utc).strftime('%Y-%m-%dT%H:%M:%SZ')}",
        "",
    ]

    if gate_failures:
        lines.append("## 闸门失败原因")
        lines.append("")
        for f in gate_failures:
            lines.append(f"- {f}")
        lines.append("")

    lines.append("## 替换明细")
    lines.append("")
    rows = _replacement_rows(results)
    if rows:
        lines.append("| 文档 | 类型 | 替换前 | 替换后 | 次数 | 建议动作 |")
        lines.append("| --- | --- | --- | --- | ---: | --- |")
        for doc_id, entity_type, original, replacement, count, suggested_action in rows:
            lines.append(
                "| "
                + " | ".join([
                    _md_cell(doc_id),
                    _md_cell(entity_type),
                    _md_cell(original),
                    _md_cell(replacement),
                    str(count),
                    _md_cell(suggested_action),
                ])
                + " |"
            )
    else:
        lines.append("无替换明细。")
    lines.append("")

    lines.append("## 各文档详情")
    lines.append("")

    for r in results:
        doc_id = r["doc_id"]
        audit_data = r.get("audit_data", {})
        summary = audit_data.get("summary", {})
        residual = audit_data.get("residual_scan", {})
        warnings = audit_data.get("warnings", [])

        lines.append(f"### {doc_id} ({r['source_ext']})")
        lines.append("")
        lines.append(f"- **Pipeline**: {r['pipeline']}")
        lines.append(f"- **Irreversible**: {r['irreversible']}")
        lines.append(f"- **实体数**: {summary.get('total_entities', 0)}")
        lines.append(f"- **替换数**: {summary.get('total_occurrences', 0)}")
        lines.append(f"- **Residual scan**: {'通过' if residual.get('passed', True) else '未通过'}")
        lines.append("")

        if warnings:
            lines.append("**Warnings:**")
            for w in warnings[:20]:
                wtype = w.get("type", "unknown")
                msg = w.get("message", "")
                text_preview = w.get("text_preview", w.get("text", ""))
                lines.append(f"- `{wtype}`: {text_preview} — {msg}")
            if len(warnings) > 20:
                lines.append(f"- ... 共 {len(warnings)} 条 warnings")
            lines.append("")

    report_path = out_dir / _REPORT_FILENAME
    report_path.write_text("\n".join(lines), encoding="utf-8")
    return report_path


# ── Manifest & source index ────────────────────────────────────────────────


def _sanitize_model_info(ner_info: dict) -> dict:
    """Keep model capability metadata in manifest without local filesystem paths."""
    if not isinstance(ner_info, dict):
        return {}

    safe = {}
    for key in ("tag_scheme", "num_labels", "id2label", "label_source", "model_io"):
        if key in ner_info:
            safe[key] = ner_info[key]
    return safe


def _write_manifest(
    results: List[dict],
    out_dir: Path,
    profile_name: str,
    allowlist_count: int,
    denylist_count: int,
    ner_info: dict,
) -> Tuple[Path, Path]:
    """Write run_manifest.json (no PII) and SENSITIVE_SOURCE_INDEX_DO_NOT_UPLOAD.json."""
    now = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")

    doc_entries = []
    source_index = {}

    for r in results:
        doc_id = r["doc_id"]
        audit_data = r.get("audit_data", {})
        summary = audit_data.get("summary", {})

        doc_entries.append({
            "document_id": doc_id,
            "source_sha256": r["source_sha256"],
            "source_ext": r["source_ext"],
            "page_count": r["page_count"],
            "pipeline": r["pipeline"],
            "irreversible": r["irreversible"],
            "entity_count": summary.get("total_entities", 0),
            "occurrence_count": summary.get("total_occurrences", 0),
            "by_entity_type": summary.get("by_entity_type", {}),
        })

        source_index[doc_id] = {
            "original_name": r["source_name"],
            "original_path": r["source_path"],
        }

    manifest = {
        "schema_version": "1.0",
        "created_at": now,
        "profile": profile_name,
        "allowlist_count": allowlist_count,
        "denylist_count": denylist_count,
        "model_info": _sanitize_model_info(ner_info),
        "document_count": len(results),
        "documents": doc_entries,
    }

    manifest_path = out_dir / _MANIFEST_FILENAME
    with open(manifest_path, "w", encoding="utf-8") as f:
        json.dump(manifest, f, ensure_ascii=False, indent=2)

    index_path = out_dir / _SOURCE_INDEX_FILENAME
    with open(index_path, "w", encoding="utf-8") as f:
        json.dump(source_index, f, ensure_ascii=False, indent=2)

    return manifest_path, index_path


# ── Cleanup ─────────────────────────────────────────────────────────────────


def _do_cleanup(
    out_dir: Path,
    cleanup: str,
    confirm_delete: bool,
    intermediate_files: List[Path],
) -> None:
    """Execute cleanup strategy.

    - none: do nothing
    - archive: move intermediate files to _archive_sensitive_do_not_upload/
    - delete: delete intermediate files (requires --confirm-delete)
    """
    if cleanup == "none":
        return

    if cleanup == "archive":
        archive_dir = out_dir / _ARCHIVE_DIR
        archive_dir.mkdir(parents=True, exist_ok=True)
        for f in intermediate_files:
            if f.exists():
                dest = archive_dir / f.name
                shutil.move(str(f), str(dest))
        return

    if cleanup == "delete":
        if not confirm_delete:
            raise BatchError(
                "--cleanup delete requires --confirm-delete flag. "
                "This operation permanently discards restore capability."
            )
        print(
            "WARNING: --cleanup delete will permanently discard map files and "
            "restore capability. This cannot be undone.",
            file=sys.stderr,
        )
        for f in intermediate_files:
            if f.exists():
                f.unlink()
        return

    raise BatchError(f"Unknown cleanup mode: {cleanup}")


# ── Main orchestrator ───────────────────────────────────────────────────────


def batch_redact_case(
    input_dir: str,
    out_dir: str,
    profile_name: str = "labor",
    allowlist_file: Optional[str] = None,
    denylist_file: Optional[str] = None,
    entity_policy_file: Optional[str] = None,
    cleanup: str = "none",
    confirm_delete: bool = False,
    model_dir: Optional[str] = None,
    rules_path: Optional[str] = None,
) -> int:
    """Run batch case redaction. Returns exit code (0 = success).

    Orchestrates: NER pre-check -> file discovery -> per-file redaction ->
    final output -> report -> gate -> cleanup -> manifest.
    """
    # ── Load config ──
    rules = load_rules(rules_path)
    profile = load_profile(profile_name, entity_policy_file=entity_policy_file)
    allowlist = load_allowlist(builtin=True, case_file=allowlist_file)
    denylist = _load_denylist(denylist_file)

    # ── Step 1: NER pre-check (fail = stop, no fallback) ──
    print("[1/8] NER pre-check...", file=sys.stderr)
    ner_info = _precheck_ner(model_dir)
    print(f"  NER OK: {ner_info.get('model_dir', 'unknown')}", file=sys.stderr)

    # ── Step 2: Discover files ──
    input_path = Path(input_dir)
    if not input_path.is_dir():
        raise BatchError(f"Input directory not found: {input_dir}")

    all_files = _discover_files(input_path)
    if not all_files:
        raise BatchError(f"No files found in input directory: {input_dir}")

    # Filter out non-document files (e.g., .DS_Store, hidden files)
    doc_files = [f for f in all_files if not f.name.startswith(".")]

    print(f"[2/8] Discovered {len(doc_files)} files", file=sys.stderr)

    # Check for scan files and verify OCR availability
    has_scan_files = any(_classify_ext(f) == "B" for f in doc_files)
    has_unsupported = any(_classify_ext(f) == "C" for f in doc_files)

    if has_unsupported:
        unsupported = [f for f in doc_files if _classify_ext(f) == "C"]
        raise BatchError(
            f"Unsupported format(s) found: {', '.join(f.suffix for f in unsupported[:5])}. "
            + _C_CONVERT_MSG.format(ext=unsupported[0].suffix)
        )

    if has_scan_files:
        _check_ocr_available()

    # ── Step 3: Process each file ──
    out_path = Path(out_dir)
    out_path.mkdir(parents=True, exist_ok=True)
    final_dir = out_path / _FINAL_DIR
    staging_final_dir = out_path / _STAGING_FINAL_DIR
    if final_dir.exists():
        raise BatchError(
            f"Final output directory already exists: {final_dir}. "
            "Use a fresh --out directory or move the existing final output first."
        )
    if staging_final_dir.exists():
        shutil.rmtree(staging_final_dir)
    staging_final_dir.mkdir(parents=True, exist_ok=True)

    results: List[dict] = []
    intermediate_files: List[Path] = []

    for idx, src in enumerate(doc_files, 1):
        doc_id = f"doc_{idx:02d}"
        print(
            f"[3/8] Processing [{idx}/{len(doc_files)}] {doc_id}: {src.name}",
            file=sys.stderr,
        )
        try:
            result = _process_one(
                src, doc_id, rules, profile, allowlist, denylist, model_dir,
            )
        except BatchError:
            raise
        except Exception as e:
            raise BatchError(f"Failed to process {src.name}: {e}")

        # Write anonymized output
        final_name = f"{doc_id}.redacted.md"
        final_path = staging_final_dir / final_name
        final_path.write_text(result["redacted_text"], encoding="utf-8")

        # Write map and audit (intermediate, may be cleaned up)
        map_path = out_path / f"{doc_id}.map.json"
        audit_path = out_path / f"{doc_id}.audit.json"

        if result["map_data"]:
            with open(map_path, "w", encoding="utf-8") as f:
                json.dump(result["map_data"], f, ensure_ascii=False, indent=2)
            intermediate_files.append(map_path)

        if result["audit_data"]:
            with open(audit_path, "w", encoding="utf-8") as f:
                json.dump(result["audit_data"], f, ensure_ascii=False, indent=2)
            intermediate_files.append(audit_path)

        results.append(result)

    print(f"[4/8] Staged final output written to {staging_final_dir}", file=sys.stderr)

    # ── Step 5: Report ──
    report_path = _generate_report(
        results, out_path, profile_name,
        len(allowlist), len(denylist),
        gate_passed=True, gate_failures=[],
    )
    print(f"[5/8] Report: {report_path.name}", file=sys.stderr)

    # ── Step 6: Validation gate ──
    print("[6/8] Running validation gate...", file=sys.stderr)
    gate_passed, gate_failures = _run_gate(results, rules, profile, denylist)

    if not gate_passed:
        # Regenerate report with gate failures
        _generate_report(
            results, out_path, profile_name,
            len(allowlist), len(denylist),
            gate_passed=False, gate_failures=gate_failures,
        )
        print(
            "GATE FAILED. Staged output preserved for debugging; final output was not created.",
            file=sys.stderr,
        )
        for f in gate_failures:
            print(f"  FAIL: {f}", file=sys.stderr)
        # Write manifest even on failure (for debugging)
        _write_manifest(results, out_path, profile_name, len(allowlist), len(denylist), ner_info)
        return 1

    print("  Gate passed.", file=sys.stderr)

    staging_final_dir.rename(final_dir)
    print(f"  Final output published to {final_dir}", file=sys.stderr)

    # ── Step 7: Cleanup (only if gate passed) ──
    print(f"[7/8] Cleanup: {cleanup}", file=sys.stderr)
    _do_cleanup(out_path, cleanup, confirm_delete, intermediate_files)

    # ── Step 8: Manifest ──
    print("[8/8] Writing manifest...", file=sys.stderr)
    _write_manifest(results, out_path, profile_name, len(allowlist), len(denylist), ner_info)

    print("Done.", file=sys.stderr)
    return 0
